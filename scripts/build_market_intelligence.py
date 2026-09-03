#!/usr/bin/env python3
"""Build public ABS market, centre, store-link, and network-history datasets."""

from __future__ import annotations

import csv
import json
import math
import os
import shutil
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - exercised by setup failures
    raise SystemExit("openpyxl is required. Run: python3 -m pip install -r requirements.txt") from exc


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
CACHE_DIR = Path(os.environ.get("ABS_CACHE_DIR", ROOT / ".cache" / "market-intelligence"))
HISTORY_DIR = DATA_DIR / "history"
STORE_CSV = DATA_DIR / "optical_stores.csv"
SA2_OUTPUT = DATA_DIR / "sa2_market.geojson"
STORE_LINK_OUTPUT = DATA_DIR / "store_market_links.json"
CENTRE_OUTPUT = DATA_DIR / "centres.json"
EVENT_OUTPUT = DATA_DIR / "network_events.json"
PROFILE_CSV = DATA_DIR / "centre_profiles.csv"
CENTRE_REGISTRY_CSV = DATA_DIR / "shopping_centres.csv"
BRAND_PROFILES = DATA_DIR / "brand_profiles.json"

ABS_RELEASE_DATE = "2026-05-26"
ABS_BASE = "https://www.abs.gov.au/methodologies/data-region-methodology/2011-25"
ABS_FILES = {
    "population": f"{ABS_BASE}/14100DO0001_2011-25.xlsx",
    "economy": f"{ABS_BASE}/14100DO0003_2011-25.xlsx",
    "income": f"{ABS_BASE}/14100DO0004_2011-25.xlsx",
    "employment": f"{ABS_BASE}/14100DO0005_2011-25.xlsx",
}
SA2_SERVICE = "https://geo.abs.gov.au/arcgis/rest/services/ASGS2021/SA2/FeatureServer/0/query"


def read_csv(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def download(url: str, target: Path) -> None:
    if target.exists() and target.stat().st_size > 100_000:
        return
    target.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "Optical leasing intelligence builder/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response, target.open("wb") as output:
        shutil.copyfileobj(response, output)


def number(value):
    if value in (None, "", "-", "np", "n.p.", "na", "n.a."):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def whole_or_decimal(value):
    if value is None:
        return None
    return int(value) if float(value).is_integer() else round(float(value), 2)


def workbook_rows(path: Path, years: set[int], columns: dict[str, tuple[int, str]]) -> dict[str, dict[int, dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Table 1"]
    for name, (index, expected_header) in columns.items():
        actual_header = str(sheet.cell(7, index + 1).value or "").strip()
        if actual_header != expected_header:
            workbook.close()
            raise RuntimeError(f"ABS workbook schema changed for {name}: expected {expected_header!r}; found {actual_header!r}")
    results: dict[str, dict[int, dict]] = defaultdict(dict)
    for row in sheet.iter_rows(min_row=8, values_only=True):
        code = str(row[0] or "").strip()
        year = int(row[2]) if isinstance(row[2], (int, float)) else None
        if len(code) != 9 or year not in years:
            continue
        results[code][year] = {name: number(row[index]) for name, (index, _) in columns.items()}
    workbook.close()
    return results


def fetch_sa2_boundaries() -> list[dict]:
    features = []
    offset = 0
    while True:
        params = {
            "where": "1=1",
            "outFields": (
                "sa2_code_2021,sa2_name_2021,state_code_2021,"
                "state_name_2021,area_albers_sqkm"
            ),
            "returnGeometry": "true",
            "outSR": "4326",
            "maxAllowableOffset": "0.01",
            "resultOffset": str(offset),
            "resultRecordCount": "2000",
            "f": "geojson",
        }
        request = urllib.request.Request(
            f"{SA2_SERVICE}?{urllib.parse.urlencode(params)}",
            headers={"User-Agent": "Optical leasing intelligence builder/1.0"},
        )
        with urllib.request.urlopen(request, timeout=120) as response:
            page = json.load(response)
        if page.get("error"):
            raise RuntimeError(f"ABS SA2 service failed: {page['error']}")
        batch = page.get("features", [])
        features.extend(batch)
        if not page.get("exceededTransferLimit") or not batch:
            break
        offset += len(batch)
    unique = {feature["properties"]["sa2_code_2021"]: feature for feature in features}
    return [unique[key] for key in sorted(unique)]


def iter_points(coordinates):
    if not coordinates:
        return
    first = coordinates[0]
    if isinstance(first, (int, float)):
        yield coordinates
    else:
        for child in coordinates:
            yield from iter_points(child)


def geometry_bounds(geometry: dict) -> tuple[float, float, float, float]:
    points = list(iter_points(geometry["coordinates"]))
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    return min(xs), min(ys), max(xs), max(ys)


def geometry_centre(geometry: dict) -> tuple[float, float]:
    min_x, min_y, max_x, max_y = geometry_bounds(geometry)
    return round((min_x + max_x) / 2, 5), round((min_y + max_y) / 2, 5)


def point_in_ring(longitude: float, latitude: float, ring: list[list[float]]) -> bool:
    inside = False
    previous = ring[-1]
    for current in ring:
        x1, y1 = previous
        x2, y2 = current
        if (y1 > latitude) != (y2 > latitude):
            boundary_x = (x2 - x1) * (latitude - y1) / ((y2 - y1) or 1e-12) + x1
            if longitude < boundary_x:
                inside = not inside
        previous = current
    return inside


def point_in_geometry(longitude: float, latitude: float, geometry: dict) -> bool:
    polygons = geometry["coordinates"]
    if geometry["type"] == "Polygon":
        polygons = [polygons]
    for polygon in polygons:
        if polygon and point_in_ring(longitude, latitude, polygon[0]):
            if not any(point_in_ring(longitude, latitude, hole) for hole in polygon[1:]):
                return True
    return False


def build_market_features() -> list[dict]:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    paths = {}
    for name, url in ABS_FILES.items():
        paths[name] = CACHE_DIR / f"abs-{name}.xlsx"
        download(url, paths[name])

    population = workbook_rows(
        paths["population"],
        {2021, 2025},
        {
            "population": (3, "Estimated resident population (no.)"),
            "density": (4, "Population density (persons/km2)"),
            "median_age": (9, "Median age - persons (years)"),
            "age_45_49_pct": (111, "Persons - 45-49 years (%)"),
            "age_50_54_pct": (112, "Persons - 50-54 years (%)"),
            "age_55_59_pct": (113, "Persons - 55-59 years (%)"),
            "age_60_64_pct": (114, "Persons - 60-64 years (%)"),
            "age_65_69_pct": (115, "Persons - 65-69 years (%)"),
            "age_70_74_pct": (116, "Persons - 70-74 years (%)"),
            "age_75_79_pct": (117, "Persons - 75-79 years (%)"),
            "age_80_84_pct": (118, "Persons - 80-84 years (%)"),
            "age_85_plus_pct": (119, "Persons - 85 and over (%)"),
        },
    )
    economy = workbook_rows(
        paths["economy"],
        {2025},
        {"total_businesses": (7, "Total businesses (no.)"), "retail_businesses": (14, "Retail trade (no.)"), "health_businesses": (24, "Health care and social assistance (no.)")},
    )
    income = workbook_rows(
        paths["income"],
        {2021},
        {"median_equivalised_household_income_weekly": (11, "Median equivalised total household income (weekly) ($)")},
    )
    employment = workbook_rows(
        paths["employment"],
        {2021},
        {"unemployment_rate": (53, "Unemployment rate (%)"), "participation_rate": (54, "Participation rate (%)")},
    )

    market_features = []
    for feature in fetch_sa2_boundaries():
        if not feature.get("geometry"):
            continue
        properties = feature["properties"]
        code = properties["sa2_code_2021"]
        current = population.get(code, {}).get(2025, {})
        baseline = population.get(code, {}).get(2021, {})
        census = population.get(code, {}).get(2021, {})
        population_current = current.get("population")
        population_baseline = baseline.get("population")
        growth = None
        if population_current is not None and population_baseline not in (None, 0):
            growth = (population_current / population_baseline - 1) * 100
        age_values = [
            census.get(key)
            for key in (
                "age_45_49_pct",
                "age_50_54_pct",
                "age_55_59_pct",
                "age_60_64_pct",
                "age_65_69_pct",
                "age_70_74_pct",
                "age_75_79_pct",
                "age_80_84_pct",
                "age_85_plus_pct",
            )
        ]
        age_45_plus = sum(value for value in age_values if value is not None) if any(
            value is not None for value in age_values
        ) else None
        small_population_caution = population_baseline is not None and population_baseline < 100
        if small_population_caution:
            growth = None
            age_45_plus = None
        if age_45_plus is not None and not 0 <= age_45_plus <= 100:
            age_45_plus = None
        unemployment = employment.get(code, {}).get(2021, {}).get("unemployment_rate")
        participation = employment.get(code, {}).get(2021, {}).get("participation_rate")
        if small_population_caution or (unemployment is not None and not 0 <= unemployment <= 100):
            unemployment = None
        if small_population_caution or (participation is not None and not 0 <= participation <= 100):
            participation = None
        equivalised_income = income.get(code, {}).get(2021, {}).get("median_equivalised_household_income_weekly")
        if small_population_caution:
            equivalised_income = None
        centre_lon, centre_lat = geometry_centre(feature["geometry"])
        market_features.append(
            {
                "type": "Feature",
                "geometry": feature["geometry"],
                "properties": {
                    "sa2_code": code,
                    "sa2_name": properties["sa2_name_2021"],
                    "state": properties["state_name_2021"],
                    "area_sqkm": whole_or_decimal(properties.get("area_albers_sqkm")),
                    "centroid_longitude": centre_lon,
                    "centroid_latitude": centre_lat,
                    "population_2025": whole_or_decimal(population_current),
                    "population_growth_2021_2025_pct": whole_or_decimal(growth),
                    "population_density_2025": whole_or_decimal(current.get("density")),
                    "median_age_2021": whole_or_decimal(census.get("median_age")),
                    "age_45_plus_pct_2021": whole_or_decimal(age_45_plus),
                    "median_equivalised_household_income_weekly_2021": whole_or_decimal(equivalised_income),
                    "unemployment_rate_2021": whole_or_decimal(unemployment),
                    "participation_rate_2021": whole_or_decimal(participation),
                    "total_businesses_2025": whole_or_decimal(
                        economy.get(code, {}).get(2025, {}).get("total_businesses")
                    ),
                    "retail_businesses_2025": whole_or_decimal(
                        economy.get(code, {}).get(2025, {}).get("retail_businesses")
                    ),
                    "health_businesses_2025": whole_or_decimal(
                        economy.get(code, {}).get(2025, {}).get("health_businesses")
                    ),
                    "source": "Australian Bureau of Statistics Data by Region 2011-25",
                    "source_url": ABS_BASE,
                    "source_release_date": ABS_RELEASE_DATE,
                    "confidence": "High",
                    "small_population_caution": small_population_caution,
                    "quality_note": "Census-derived rates and growth suppressed because the 2021 population was below 100" if small_population_caution else "Published ABS indicators; Census measures retain their 2021 reference year",
                },
            }
        )
    return market_features


def query_sa2_for_point(longitude: float, latitude: float) -> dict:
    params = {
        "where": "1=1",
        "geometry": f"{longitude},{latitude}",
        "geometryType": "esriGeometryPoint",
        "inSR": "4326",
        "spatialRel": "esriSpatialRelIntersects",
        "outFields": "sa2_code_2021,sa2_name_2021",
        "returnGeometry": "false",
        "f": "json",
    }
    request = urllib.request.Request(
        f"{SA2_SERVICE}?{urllib.parse.urlencode(params)}",
        headers={"User-Agent": "Optical leasing intelligence builder/1.0"},
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        payload = json.load(response)
    matches = payload.get("features", [])
    return matches[0]["attributes"] if matches else {}


def point_to_segment_km(
    longitude: float, latitude: float, first: list[float], second: list[float]
) -> float:
    scale_x = 111.32 * math.cos(math.radians(latitude))
    ax, ay = (first[0] - longitude) * scale_x, (first[1] - latitude) * 110.574
    bx, by = (second[0] - longitude) * scale_x, (second[1] - latitude) * 110.574
    dx, dy = bx - ax, by - ay
    denominator = dx * dx + dy * dy
    fraction = 0.0 if denominator == 0 else max(0.0, min(1.0, -(ax * dx + ay * dy) / denominator))
    return math.hypot(ax + fraction * dx, ay + fraction * dy)


def geometry_boundary_distance_km(longitude: float, latitude: float, geometry: dict) -> float:
    polygons = geometry.get("coordinates", [])
    if geometry.get("type") == "Polygon":
        polygons = [polygons]
    distances = []
    for polygon in polygons:
        if not polygon:
            continue
        ring = polygon[0]
        distances.extend(
            point_to_segment_km(longitude, latitude, ring[index], ring[index + 1])
            for index in range(len(ring) - 1)
        )
    return min(distances, default=float("inf"))


def link_stores_to_market(stores: list[dict], features: list[dict]) -> dict[str, dict]:
    spatial = []
    for feature in features:
        spatial.append((geometry_bounds(feature["geometry"]), feature))
    links = {}
    for store in stores:
        if store.get("country") != "Australia":
            links[store["store_id"]] = {
                "sa2_code": "",
                "sa2_name": "",
                "match_confidence": "Not available",
                "geography_system": "Stats NZ",
                "coverage_note": "New Zealand demographic catchment metrics are not yet published in this build.",
            }
            continue
        longitude = float(store["longitude"])
        latitude = float(store["latitude"])
        match = None
        for bounds, feature in spatial:
            min_x, min_y, max_x, max_y = bounds
            if min_x <= longitude <= max_x and min_y <= latitude <= max_y:
                if point_in_geometry(longitude, latitude, feature["geometry"]):
                    match = feature
                    break
        exact = {}
        if not match:
            exact = query_sa2_for_point(longitude, latitude)
        boundary_match = None
        boundary_distance = None
        if not match and not exact:
            nearest = min(
                (
                    (geometry_boundary_distance_km(longitude, latitude, feature["geometry"]), feature)
                    for feature in features
                ),
                key=lambda item: item[0],
            )
            if nearest[0] <= 2.0:
                boundary_distance, boundary_match = nearest
        links[store["store_id"]] = {
            "sa2_code": (
                match["properties"]["sa2_code"]
                if match
                else exact.get("sa2_code_2021", "")
                or (boundary_match or {}).get("properties", {}).get("sa2_code", "")
            ),
            "sa2_name": (
                match["properties"]["sa2_name"]
                if match
                else exact.get("sa2_name_2021", "")
                or (boundary_match or {}).get("properties", {}).get("sa2_name", "")
            ),
            "match_confidence": "High" if match or exact else "Medium" if boundary_match else "Unmatched",
            "match_method": "Point in SA2" if match or exact else "Nearest SA2 boundary fallback" if boundary_match else "Unmatched",
            "boundary_distance_km": round(boundary_distance, 3) if boundary_distance is not None else None,
        }
    return links


def build_centres(stores: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for store in stores:
        if store.get("venue_id"):
            grouped[store["venue_id"]].append(store)
    overrides = {row["venue_id"]: row for row in read_csv(PROFILE_CSV)}
    registry = {row["venue_id"]: row for row in read_csv(CENTRE_REGISTRY_CSV)}
    centres = []
    for venue_id, members in sorted(grouped.items()):
        override = overrides.get(venue_id, {})
        registered = registry.get(venue_id, {})
        venue_name = registered.get("venue_name", "") or members[0]["venue_name"]
        manager = override.get("manager", "")
        public_url = override.get("public_url", "") or registered.get("official_url", "")
        confidence = override.get("confidence", "") or registered.get("confidence", "")
        if not manager and "westfield" in venue_name.lower():
            manager = "Scentre Group"
            public_url = "https://www.westfield.com.au/"
            confidence = "Medium"
        if not manager and "stockland" in venue_name.lower():
            manager = "Stockland"
            public_url = "https://www.stockland.com.au/shopping-centres"
            confidence = "Medium"
        centres.append(
            {
                "centre_id": venue_id,
                "name": venue_name,
                "country": members[0].get("country", "Australia"),
                "state": registered.get("state", "") or members[0]["state"],
                "suburb": registered.get("suburb", "") or members[0]["suburb"],
                "latitude": round(
                    float(registered["latitude"])
                    if registered.get("latitude")
                    else sum(float(item["latitude"]) for item in members) / len(members),
                    6,
                ),
                "longitude": round(
                    float(registered["longitude"])
                    if registered.get("longitude")
                    else sum(float(item["longitude"]) for item in members) / len(members),
                    6,
                ),
                "retailers": sorted({item["retailer"] for item in members}),
                "optical_store_count": len(members),
                "owner": override.get("owner", ""),
                "manager": manager,
                "centre_type": override.get("centre_type", ""),
                "gla_sqm": whole_or_decimal(number(override.get("gla_sqm"))),
                "annual_visits": whole_or_decimal(number(override.get("annual_visits"))),
                "trade_area_population": whole_or_decimal(number(override.get("trade_area_population"))),
                "anchors": [item.strip() for item in override.get("anchors", "").split(";") if item.strip()],
                "tenancy_count": whole_or_decimal(number(override.get("tenancy_count"))),
                "redevelopment_activity": override.get("redevelopment_activity", ""),
                "leasing_contact": override.get("leasing_contact", ""),
                "public_url": public_url,
                "metrics_date": override.get("metrics_date", ""),
                "confidence": confidence or "Base",
                "source_basis": (
                    "Curated public centre profile"
                    if override
                    else (
                        "Verified public shopping-centre registry"
                        if registered
                        else "Centre entity derived from reviewed store venue IDs"
                    )
                ),
            }
        )
    return centres


def haversine(first: dict, second: dict) -> float:
    radius = 6371.0088
    lat1 = math.radians(float(first["latitude"]))
    lat2 = math.radians(float(second["latitude"]))
    delta_lat = math.radians(float(second["latitude"]) - float(first["latitude"]))
    delta_lon = math.radians(float(second["longitude"]) - float(first["longitude"]))
    value = (
        math.sin(delta_lat / 2) ** 2
        + math.cos(lat1) * math.cos(lat2) * math.sin(delta_lon / 2) ** 2
    )
    return radius * 2 * math.atan2(math.sqrt(value), math.sqrt(1 - value))


def build_history(stores: list[dict]) -> dict:
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    snapshot_date = max(str(store["fetched_at"])[:10] for store in stores)
    snapshot_path = HISTORY_DIR / f"{snapshot_date}.json"
    current = {
        store["store_id"]: {
            key: store.get(key, "")
            for key in (
                "store_id",
                "retailer",
                "country",
                "name",
                "status",
                "state",
                "suburb",
                "full_address",
                "latitude",
                "longitude",
                "venue_id",
            )
        }
        for store in stores
    }
    prior_paths = [path for path in sorted(HISTORY_DIR.glob("*.json")) if path != snapshot_path]
    health_path = DATA_DIR / "data_health.json"
    health = json.loads(health_path.read_text(encoding="utf-8")) if health_path.exists() else {}
    if health.get("certification_status") != "Certified" and os.environ.get("ALLOW_UNCERTIFIED_SNAPSHOT") != "1":
        latest_path = sorted(HISTORY_DIR.glob("*.json"))[-1] if list(HISTORY_DIR.glob("*.json")) else None
        baseline_date = json.loads(latest_path.read_text(encoding="utf-8"))["snapshot_date"] if latest_path else snapshot_date
        return {
            "baseline_date": baseline_date,
            "current_snapshot_date": snapshot_date,
            "event_count": 0,
            "events": [],
            "coverage_baselines_added": [],
            "archived": False,
            "note": "Snapshot not archived because the current census has not passed data-health certification.",
        }
    events = []
    coverage_baselines_added = []
    baseline_date = snapshot_date
    if prior_paths:
        prior_payload = json.loads(prior_paths[-1].read_text(encoding="utf-8"))
        prior = prior_payload["stores"]
        baseline_date = prior_payload["snapshot_date"]
        prior_scopes = {
            (store["retailer"], store.get("country", "Australia")) for store in prior.values()
        }
        current_scopes = {(store["retailer"], store["country"]) for store in current.values()}
        coverage_baselines_added = [
            {"retailer": retailer, "country": country}
            for retailer, country in sorted(current_scopes - prior_scopes)
        ]
        for store_id in sorted(current.keys() - prior.keys()):
            store = current[store_id]
            if (store["retailer"], store["country"]) in prior_scopes:
                events.append({"type": "Opened", "date": snapshot_date, **store})
        for store_id in sorted(prior.keys() - current.keys()):
            store = prior[store_id]
            scope = (store["retailer"], store.get("country", "Australia"))
            if scope in current_scopes:
                events.append({"type": "Closed", "date": snapshot_date, **store})
        for store_id in sorted(current.keys() & prior.keys()):
            before = prior[store_id]
            after = current[store_id]
            moved = haversine(before, after) >= 0.1 or before["full_address"] != after["full_address"]
            if moved:
                events.append(
                    {
                        "type": "Relocated",
                        "date": snapshot_date,
                        **after,
                        "previous_address": before["full_address"],
                        "distance_moved_km": round(haversine(before, after), 2),
                    }
                )
    snapshot_path.write_text(
        json.dumps({"snapshot_date": snapshot_date, "stores": current}, indent=2) + "\n",
        encoding="utf-8",
    )
    return {
        "baseline_date": baseline_date,
        "current_snapshot_date": snapshot_date,
        "event_count": len(events),
        "events": events,
        "coverage_baselines_added": coverage_baselines_added,
        "note": (
            "This is the first archived baseline; changes will appear after the next successful refresh."
            if not prior_paths
            else (
                "New retailer or country coverage was added as a baseline, not counted as store openings."
                if coverage_baselines_added
                else "Events compare the current network with the latest prior successful snapshot."
            )
        ),
    }


def main() -> None:
    if not STORE_CSV.exists() or not BRAND_PROFILES.exists():
        raise SystemExit("Run scripts/build_optical_network.py before building market intelligence.")
    stores = read_csv(STORE_CSV)
    features = build_market_features()
    links = link_stores_to_market(stores, features)
    centres = build_centres(stores)
    history = build_history(stores)

    SA2_OUTPUT.write_text(
        json.dumps(
            {
                "type": "FeatureCollection",
                "metadata": {
                    "name": "ABS SA2 leasing market indicators",
                    "feature_count": len(features),
                    "source": "Australian Bureau of Statistics Data by Region 2011-25",
                    "source_url": ABS_BASE,
                    "source_release_date": ABS_RELEASE_DATE,
                    "boundary_source": SA2_SERVICE.rsplit("/query", 1)[0],
                    "field_definitions": {
                        "population_2025": "ABS estimated resident population",
                        "population_growth_2021_2025_pct": "Change in ABS estimated resident population; suppressed where the 2021 population is below 100",
                        "median_equivalised_household_income_weekly_2021": "ABS Census median equivalised total household income per week; this is not consumer spending",
                    },
                    "consumer_spending_status": "Not included — no verified consumer-expenditure dataset is loaded",
                    "generated_at": datetime.now(timezone.utc).isoformat(),
                },
                "features": features,
            },
            separators=(",", ":"),
        )
        + "\n",
        encoding="utf-8",
    )
    STORE_LINK_OUTPUT.write_text(
        json.dumps(
            {
                "metadata": {
                    "store_count": len(stores),
                    "matched_count": sum(bool(link["sa2_code"]) for link in links.values()),
                    "source_release_date": ABS_RELEASE_DATE,
                },
                "links": links,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    CENTRE_OUTPUT.write_text(
        json.dumps(
            {
                "metadata": {
                    "centre_count": len(centres),
                    "enriched_count": sum(centre["confidence"] == "High" for centre in centres),
                    "generated_at": datetime.now(timezone.utc).isoformat(),
                },
                "centres": centres,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    EVENT_OUTPUT.write_text(json.dumps(history, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(features)} SA2 market features to {SA2_OUTPUT}")
    print(f"Matched {sum(bool(link['sa2_code']) for link in links.values())}/{len(stores)} stores to SA2")
    print(f"Wrote {len(centres)} reviewed centre entities")
    if history.get("archived", True):
        print(f"Archived network snapshot with {history['event_count']} change events")
    else:
        print("Network snapshot was not archived because census certification is incomplete")


if __name__ == "__main__":
    main()
